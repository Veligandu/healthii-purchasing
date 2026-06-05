"""
Healthii Purchasing Agent – Leopold
Wöchentlicher Bestellvorschlag mit Google Drive Archivierung.

Verwendung:
  1. wiederbestellung.xlsx in diesen Ordner legen
  2. Skript starten:
     python3 purchasing_agent.py

  Beim ersten Start: python3 purchasing_agent.py --auth

Ausgabe:
  - Purchase-Order-KWXX-YYYY.xlsx  (lokal im Projektordner)
  - Archiv in Leopolds Google Drive:
      Healthii Purchasing/YYYY/KWXX/

Abhängigkeiten (pip install):
  pandas openpyxl
  google-api-python-client google-auth-httplib2 google-auth-oauthlib
"""

import argparse
import base64
import glob
import io
import math
import os
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ─── Konfiguration ────────────────────────────────────────────────────────────

SCOPES = [
    'https://www.googleapis.com/auth/gmail.modify',
    'https://www.googleapis.com/auth/drive',
]

BASE_DIR          = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_PATH  = os.path.join(BASE_DIR, 'credentials.json')
TOKEN_PATH        = os.path.join(BASE_DIR, 'token.json')

FRANK_EMAIL           = 'frank.mueller@healthii.de'
DRIVE_ROOT            = 'Healthii Purchasing'
WIEDERBESTELLUNG_PATH = os.path.join(BASE_DIR, 'wiederbestellung.xlsx')

CONFIG = {
    'gewichtung_l30':            0.7,
    'gewichtung_l90':            0.3,
    'ziel_tage':                 60,
    'mbw_standard':              2000.0,
    'kritische_positionsgroesse': 0.0,   # 0 = deaktiviert
    'mindestreichweite':          30,
}

# ─── Auth ─────────────────────────────────────────────────────────────────────

def get_services():
    """Gibt authentifizierte Gmail- und Drive-Service-Objekte zurück.
    Unterstützt lokalen Modus (token.json) und Cloud-Modus (Streamlit Secrets).
    """
    creds = None

    # Cloud-Modus: Token aus Streamlit Secrets
    try:
        import streamlit as st
        import json as _json
        if 'GOOGLE_TOKEN' in st.secrets:
            creds = Credentials.from_authorized_user_info(
                _json.loads(st.secrets['GOOGLE_TOKEN']), SCOPES
            )
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
    except Exception:
        pass

    # Lokaler Modus: Token aus Datei
    if creds is None:
        if os.path.exists(TOKEN_PATH):
            creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(TOKEN_PATH, 'w') as f:
                f.write(creds.to_json())

    gmail = build('gmail', 'v1', credentials=creds)
    drive = build('drive', 'v3', credentials=creds)
    return gmail, drive

# ─── Drive Helpers ─────────────────────────────────────────────────────────────

def get_or_create_folder(drive, name, parent_id=None):
    """Gibt die ID eines Ordners zurück, erstellt ihn falls nicht vorhanden."""
    query = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        query += f" and '{parent_id}' in parents"
    results = drive.files().list(q=query, fields='files(id)').execute()
    files = results.get('files', [])
    if files:
        return files[0]['id']
    metadata = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        metadata['parents'] = [parent_id]
    return drive.files().create(body=metadata, fields='id').execute()['id']


def get_week_folder_id(drive, kw, year):
    """Erstellt/gibt KW-Ordner zurück: Healthii Purchasing/YYYY/KWXX"""
    root_id = get_or_create_folder(drive, DRIVE_ROOT)
    year_id = get_or_create_folder(drive, str(year), root_id)
    return get_or_create_folder(drive, f'KW{kw:02d}', year_id)


def get_stammdaten_folder_id(drive):
    root_id = get_or_create_folder(drive, DRIVE_ROOT)
    return get_or_create_folder(drive, 'Stammdaten', root_id)


def upload_bytes_to_drive(drive, data: bytes, filename: str, folder_id: str, mimetype: str):
    """Lädt Bytes in Drive hoch; überschreibt falls Datei bereits existiert."""
    query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
    existing = drive.files().list(q=query, fields='files(id)').execute().get('files', [])
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mimetype, resumable=False)
    if existing:
        drive.files().update(fileId=existing[0]['id'], media_body=media).execute()
    else:
        drive.files().create(
            body={'name': filename, 'parents': [folder_id]},
            media_body=media, fields='id',
        ).execute()
    print(f"Drive: {filename}")


def download_csv_from_drive(drive, filename: str, folder_id: str):
    """Lädt CSV aus Drive als DataFrame; gibt None zurück falls nicht vorhanden."""
    query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
    files = drive.files().list(q=query, fields='files(id)').execute().get('files', [])
    if not files:
        return None
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, drive.files().get_media(fileId=files[0]['id']))
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return pd.read_csv(buf)



# ─── Gmail Helpers ─────────────────────────────────────────────────────────────



def send_order_email(gmail, hersteller, df_gruppe, kw, year, bestellwert, mbw):
    """Sendet vorbereitete Bestell-E-Mail für einen Hersteller an Frank."""
    rows_html = ''
    for _, row in df_gruppe.iterrows():
        marge = row.get('Marge_EK_AEP')
        marge_str = f"{marge:.1%}" if marge is not None and not pd.isna(marge) else '–'
        rows_html += f"""
        <tr>
          <td style="padding:4px 8px;border-bottom:1px solid #eee;">{row['Pzn']}</td>
          <td style="padding:4px 8px;border-bottom:1px solid #eee;">{row['Artikelname']}</td>
          <td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:center;">{int(row['Bestellmenge'])}</td>
          <td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:right;">{row['Rechnungs Netto Ek Ve1']:.2f} €</td>
          <td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:right;">{row.get('Aep', 0):.2f} €</td>
          <td style="padding:4px 8px;border-bottom:1px solid #eee;text-align:right;">{marge_str}</td>
        </tr>"""

    html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#2c2c2a;font-size:13px;">
      <p>Hallo Frank,</p>
      <p>hier ist der Bestellvorschlag für <strong>{hersteller}</strong> (KW{kw:02d}/{year}):</p>
      <table style="border-collapse:collapse;width:100%;">
        <thead>
          <tr style="background:#2c2c2a;color:white;">
            <th style="padding:6px 8px;text-align:left;">PZN</th>
            <th style="padding:6px 8px;text-align:left;">Artikelname</th>
            <th style="padding:6px 8px;text-align:center;">Bestellmenge</th>
            <th style="padding:6px 8px;text-align:right;">EK je Einheit</th>
            <th style="padding:6px 8px;text-align:right;">AEP</th>
            <th style="padding:6px 8px;text-align:right;">1-EK/AEP</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
        <tfoot>
          <tr style="background:#1D9E75;color:white;font-weight:bold;">
            <td colspan="3" style="padding:6px 8px;">Summe {hersteller}</td>
            <td colspan="2" style="padding:6px 8px;text-align:right;">Bestellwert:</td>
            <td style="padding:6px 8px;text-align:right;">{bestellwert:,.2f} €</td>
          </tr>
        </tfoot>
      </table>
      <p style="color:#888;font-size:11px;margin-top:16px;">
        MBW: {mbw:.0f} € — Dies ist ein Bestellvorschlag, keine Bestellung.
      </p>
      <p>Leopold</p>
    </body></html>"""

    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'[Leopold] Bestellvorschlag KW{kw:02d} – {hersteller}'
    msg['To'] = FRANK_EMAIL
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    gmail.users().messages().send(userId='me', body={'raw': raw}).execute()
    print(f"  E-Mail: {hersteller} ({bestellwert:,.2f} €)")


# ─── Bestellhistorie ──────────────────────────────────────────────────────────

def finde_letzte_bestellung_excel():
    """Gibt den Pfad zur neuesten bestellhistorie-*.xlsx zurück — nie die heutige Ausgabedatei."""
    dateien = glob.glob(os.path.join(BASE_DIR, 'bestellhistorie-*.xlsx'))
    if not dateien:
        return None
    # Heutige Datei ausschließen (wäre Ausgabe des laufenden Prozesses)
    heute = date.today().strftime('%d%m%Y')
    kandidaten = [f for f in dateien if heute not in os.path.basename(f)]
    if not kandidaten:
        kandidaten = dateien  # Fallback: nur heutige vorhanden (erster Lauf überhaupt)
    return max(kandidaten, key=os.path.getmtime)


def lade_letzte_bestellung(pfad):
    """Liest nicht-eingelagerte Bestellmengen aus dem letzten Purchase-Order Excel."""
    if pfad is None:
        return None
    df = pd.read_excel(pfad)
    df['Pzn'] = df['Pzn'].astype(str)
    nicht_eingelagert = df[df['eingelagert'].str.strip().str.lower() == 'nein'].copy()
    if nicht_eingelagert.empty:
        return None
    print(f"Letzte Bestellung: {os.path.basename(pfad)} — "
          f"{len(nicht_eingelagert)} Positionen noch nicht eingelagert")
    return nicht_eingelagert[['Pzn', 'Bestellmenge']].rename(
        columns={'Bestellmenge': 'Bestellmenge_letzte_Woche'}
    )


def speichere_bestellhistorie(df_input, df_bestellen):
    """Speichert die neue Bestellhistorie im purchase-orders Format."""
    bestellte_pzn   = set(df_bestellen['Pzn'].astype(str))
    bestellmengen   = df_bestellen.set_index('Pzn')['Bestellmenge'].to_dict()

    df_out = df_input[df_input['Pzn'].astype(str).isin(bestellte_pzn)].copy()
    df_out['Bestellmenge'] = df_out['Pzn'].astype(str).map(bestellmengen)
    df_out['eingelagert']  = 'nein'

    filename = f"bestellhistorie-{date.today().strftime('%d%m%Y')}.xlsx"
    pfad     = os.path.join(BASE_DIR, filename)
    df_out.to_excel(pfad, index=False, sheet_name='Abfrageergebnis')
    print(f"Bestellhistorie gespeichert: {filename}")
    return filename, pfad


# ─── Hilfsfunktionen ──────────────────────────────────────────────────────────

def aufrunden_ve(bedarf: float, ve: int) -> int:
    if bedarf <= 0:
        return 0
    return math.ceil(bedarf / ve) * ve

# ─── Kernlogik ────────────────────────────────────────────────────────────────

def berechne_bestellvorschlag(excel_bytes: bytes, letzte_bestellung_df, mbw_ausnahmen: dict) -> dict:
    log = []
    df = pd.read_excel(io.BytesIO(excel_bytes))
    df['Pzn'] = df['Pzn'].astype(str)
    df_input = df.copy()  # Original für Bestellhistorie

    log.append(f"Artikel geladen: {len(df)}, Hersteller: {df['Hersteller'].nunique()}")

    w30 = CONFIG['gewichtung_l30']
    w90 = CONFIG['gewichtung_l90']
    df['TV']         = (w30 * df['Verkaeufe L30'] / 30) + (w90 * df['Verkaeufe L90'] / 90)
    df['Ziel_Menge'] = (df['TV'] * CONFIG['ziel_tage']).round(0)
    df['Lagerbestand'] = df['Lagerbestand'].fillna(0)
    df['Ve1']          = df['Ve1'].fillna(1).astype(int)

    if letzte_bestellung_df is not None:
        letzte_bestellung_df['Pzn'] = letzte_bestellung_df['Pzn'].astype(str)
        df = df.merge(letzte_bestellung_df, on='Pzn', how='left')
    else:
        df['Bestellmenge_letzte_Woche'] = 0

    df['Bestellmenge_letzte_Woche'] = df['Bestellmenge_letzte_Woche'].fillna(0)
    df['Effektiver_Bestand'] = df['Lagerbestand'] + df['Bestellmenge_letzte_Woche']
    df['Bedarf_roh']   = df['Ziel_Menge'] - df['Effektiver_Bestand']
    df['Bestellmenge'] = df.apply(lambda r: aufrunden_ve(r['Bedarf_roh'], int(r['Ve1'])), axis=1)

    # Kritische Positionsgröße: Bestellwert auf Minimum für Mindestreichweite kürzen
    krit_eur  = CONFIG.get('kritische_positionsgroesse', 0.0)
    mind_tage = CONFIG.get('mindestreichweite', 30)
    if krit_eur > 0:
        def _reduziere(row):
            bm  = row['Bestellmenge']
            ek  = row.get('Rechnungs Netto Ek Ve1', 0) or 0
            ve1 = max(1, int(row['Ve1']))
            tv  = row['TV']
            if bm * ek <= krit_eur:
                return bm   # unter Grenze — nichts tun
            if tv <= 0:
                return bm   # kein Verbrauch — nichts tun
            # Mindestmenge für Mindestreichweite berechnen (aufgerundet auf Ve1)
            min_roh  = max(0, mind_tage * tv - row['Effektiver_Bestand'])
            min_menge = aufrunden_ve(min_roh, ve1)
            # Mindestmenge verwenden (kann immer noch über Grenze liegen — ist dann unvermeidbar)
            return min_menge
        df['Bestellmenge'] = df.apply(_reduziere, axis=1)
        log.append(f"⚙ Kritische Positionsgröße aktiv: >{krit_eur:.0f} € → Minimum für {mind_tage} Tage Reichweite")

    # Bestellwert für alle Artikel (auch kein Bedarf) vorberechnen
    df['Bestellwert'] = df['Bestellmenge'] * df['Rechnungs Netto Ek Ve1']

    kandidaten = df[df['Bestellmenge'] > 0].copy()

    bestellen_liste  = []
    unter_mbw_liste  = []
    hersteller_log   = {}   # strukturiert: {hersteller: {status, mbw, bestellwert, df_positionen}}

    # Alle Hersteller aus Gesamtliste (inkl. kein Bedarf)
    for hersteller, grp_all in df.groupby('Hersteller'):
        mbw = mbw_ausnahmen.get(hersteller, CONFIG['mbw_standard'])
        grp_all = grp_all.copy()
        grp_all['MBW'] = mbw

        grp_kand = grp_all[grp_all['Bestellmenge'] > 0].copy()
        gesamtwert = grp_kand['Bestellwert'].sum() if not grp_kand.empty else 0.0

        if not grp_kand.empty:
            grp_kand['Fehlbetrag'] = max(0, mbw - gesamtwert)
            if gesamtwert >= mbw:
                status = 'bestellen'
                bestellen_liste.append(grp_kand)
                log.append(f"✓ BESTELLEN: {hersteller} — {len(grp_kand)} Pos., {gesamtwert:.2f} EUR")
            else:
                status = 'unter_mbw'
                unter_mbw_liste.append(grp_kand)
                log.append(f"✗ UNTER MBW: {hersteller} — {gesamtwert:.2f} EUR (fehlt: {mbw - gesamtwert:.2f} EUR)")
        else:
            status = 'kein_bedarf'
            log.append(f"— KEIN BEDARF: {hersteller} — Lagerbestand ausreichend")

        hersteller_log[hersteller] = {
            'status':      status,
            'mbw':         mbw,
            'bestellwert': gesamtwert,
            'fehlbetrag':  max(0, mbw - gesamtwert),
            'df':          grp_all[[
                'Pzn', 'Artikelname', 'Lagerbestand',
                'Bestellmenge_letzte_Woche', 'Effektiver_Bestand',
                'Verkaeufe L30', 'Verkaeufe L90', 'TV',
                'Ziel_Menge', 'Bedarf_roh', 'Ve1',
                'Bestellmenge', 'Rechnungs Netto Ek Ve1', 'Bestellwert',
            ]].reset_index(drop=True),
        }

    return {
        'bestellen':     pd.concat(bestellen_liste) if bestellen_liste else pd.DataFrame(),
        'unter_mbw':     pd.concat(unter_mbw_liste) if unter_mbw_liste else pd.DataFrame(),
        'log':           log,
        'df_input':      df_input,
        'hersteller_log': hersteller_log,
    }

# ─── Excel Output ─────────────────────────────────────────────────────────────

HEALTHII_GRUEN = '1D9E75'
HEALTHII_HELL  = 'E1F5EE'
GRAU_KOPF      = '2C2C2A'
WEISS          = 'FFFFFF'


def erstelle_bestellsheet(ergebnis: dict, kw: int, year: int) -> bytes:
    """Erstellt das Purchase-Order Excel und gibt es als Bytes zurück."""
    df       = ergebnis['bestellen']
    unter_mbw = ergebnis['unter_mbw']
    if df.empty:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bestellvorschläge'

    def berechne_marge(row):
        ek  = row.get('Rechnungs Netto Ek Ve1')
        aep = row.get('Aep')
        if pd.isna(ek) or pd.isna(aep) or aep == 0:
            return None
        return 1 - ek / aep

    df = df.copy()
    df['Marge_EK_AEP'] = df.apply(berechne_marge, axis=1)

    spalten = [
        ('PZN',               'Pzn',                    12),
        ('Artikelname',       'Artikelname',             45),
        ('Hersteller',        'Hersteller',              32),
        ('Bestellmenge',      'Bestellmenge',            13),
        ('EK je Einheit (€)', 'Rechnungs Netto Ek Ve1',  16),
        ('Bestellwert (€)',   'Bestellwert',             15),
        ('AEP (€)',           'Aep',                     14),
        ('1-EK/AEP',          'Marge_EK_AEP',            12),
        ('Lagerbestand',      'Lagerbestand',            13),
        ('L30',               'Verkaeufe L30',           10),
        ('L90',               'Verkaeufe L90',           10),
        ('Ve2',               'Ve2',                      8),
        ('EK Ve2 (€)',        'Rechnungs Netto Ek Ve2',  14),
    ]

    header_font = Font(bold=True, color=WEISS, name='Arial', size=10)
    header_fill = PatternFill('solid', fgColor=GRAU_KOPF)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='EEEEEE'),
    )

    ws.row_dimensions[1].height = 32
    for col_idx, (header, _, width) in enumerate(spalten, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    hersteller_farben = {}
    farb_pool = [HEALTHII_HELL, 'F5F5F5']
    farb_idx  = 0
    row = 2

    for hersteller, grp in df.groupby('Hersteller'):
        if hersteller not in hersteller_farben:
            hersteller_farben[hersteller] = farb_pool[farb_idx % 2]
            farb_idx += 1
        fill = PatternFill('solid', fgColor=hersteller_farben[hersteller])

        for _, zeile in grp.iterrows():
            for col_idx, (_, feld, _) in enumerate(spalten, 1):
                wert = zeile.get(feld, '')
                if pd.isna(wert):
                    wert = ''
                cell = ws.cell(row=row, column=col_idx, value=wert)
                cell.fill   = fill
                cell.border = thin_border
                cell.font   = Font(name='Arial', size=10)
                if col_idx in (4, 9, 10, 11, 12):
                    cell.alignment = Alignment(horizontal='center')
                if col_idx in (5, 6, 7):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                if col_idx == 8:
                    cell.number_format = '0.0%'
                    cell.alignment = Alignment(horizontal='right')
                if col_idx == 13:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    # Grün hervorheben wenn EK Ve2 günstiger als EK Ve1
                    ek1 = zeile.get('Rechnungs Netto Ek Ve1')
                    ek2 = zeile.get('Rechnungs Netto Ek Ve2')
                    if pd.notna(ek1) and pd.notna(ek2) and ek2 < ek1:
                        cell.fill = PatternFill('solid', fgColor='27AE60')
                        cell.font = Font(name='Arial', size=10, bold=True, color=WEISS)
            row += 1

        summen_fill = PatternFill('solid', fgColor=HEALTHII_GRUEN)
        summen_font = Font(bold=True, color=WEISS, name='Arial', size=10)
        mbw_val = grp['MBW'].iloc[0] if 'MBW' in grp.columns else CONFIG.get('mbw_standard', 2000.0)
        for col_idx in range(1, len(spalten) + 1):
            cell = ws.cell(row=row, column=col_idx, value='')
            cell.fill = summen_fill
            cell.font = summen_font
        ws.cell(row=row, column=2, value=f'Summe {hersteller}').font = summen_font
        ws.cell(row=row, column=2).fill = summen_fill
        ws.cell(row=row, column=4, value=int(grp['Bestellmenge'].sum()))
        ws.cell(row=row, column=4).font      = summen_font
        ws.cell(row=row, column=4).fill      = summen_fill
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6, value=round(grp['Bestellwert'].sum(), 2))
        ws.cell(row=row, column=6).font         = summen_font
        ws.cell(row=row, column=6).fill         = summen_fill
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=6).alignment    = Alignment(horizontal='right')
        ws.cell(row=row, column=3, value=f'MBW: {mbw_val:.0f} €').font = Font(italic=True, color=WEISS, name='Arial', size=9)
        ws.cell(row=row, column=3).fill = summen_fill
        row += 2

    gesamt_total = df['Bestellwert'].sum()
    ws.insert_rows(1)
    ws.merge_cells('A1:M1')
    info = ws['A1']
    info.value = (
        f"Bestellvorschlag KW{kw:02d}/{year}  |  "
        f"Gesamtbestellwert: {gesamt_total:,.2f} EUR  |  "
        f"Hersteller: {df['Hersteller'].nunique()}  |  "
        f"Positionen: {len(df)}"
    )
    info.font      = Font(bold=True, color=WEISS, name='Arial', size=11)
    info.fill      = PatternFill('solid', fgColor=HEALTHII_GRUEN)
    info.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Tab 2: Unter MBW – gleiche Optik wie Tab 1 ──
    ws2 = wb.create_sheet('Unter MBW – nicht bestellt')

    if not unter_mbw.empty:
        unter_mbw = unter_mbw.copy()
        unter_mbw['Marge_EK_AEP'] = unter_mbw.apply(berechne_marge, axis=1)

        ws2.row_dimensions[1].height = 32
        for col_idx, (header, _, width) in enumerate(spalten, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

        hersteller_farben2 = {}
        farb_idx2 = 0
        row2 = 2

        for hersteller, grp in unter_mbw.groupby('Hersteller'):
            if hersteller not in hersteller_farben2:
                hersteller_farben2[hersteller] = farb_pool[farb_idx2 % 2]
                farb_idx2 += 1
            fill2 = PatternFill('solid', fgColor=hersteller_farben2[hersteller])

            for _, zeile in grp.iterrows():
                for col_idx, (_, feld, _) in enumerate(spalten, 1):
                    wert = zeile.get(feld, '')
                    if pd.isna(wert):
                        wert = ''
                    cell = ws2.cell(row=row2, column=col_idx, value=wert)
                    cell.fill   = fill2
                    cell.border = thin_border
                    cell.font   = Font(name='Arial', size=10)
                    if col_idx in (4, 9, 10, 11):
                        cell.alignment = Alignment(horizontal='center')
                    if col_idx in (5, 6, 7):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    if col_idx == 8:
                        cell.number_format = '0.0%'
                        cell.alignment = Alignment(horizontal='right')
                row2 += 1

            # Summenzeile mit MBW- und Fehlbetrag-Hinweis
            mbw_val    = grp['MBW'].iloc[0] if 'MBW' in grp.columns else CONFIG.get('mbw_standard', 2000.0)
            fehlbetrag = grp['Fehlbetrag'].iloc[0] if 'Fehlbetrag' in grp.columns else 0.0
            summen_fill2 = PatternFill('solid', fgColor='C0392B')   # Rot für Unter-MBW
            summen_font2 = Font(bold=True, color=WEISS, name='Arial', size=10)
            for col_idx in range(1, len(spalten) + 1):
                cell = ws2.cell(row=row2, column=col_idx, value='')
                cell.fill = summen_fill2
                cell.font = summen_font2
            ws2.cell(row=row2, column=2, value=f'Summe {hersteller}').font = summen_font2
            ws2.cell(row=row2, column=2).fill = summen_fill2
            ws2.cell(row=row2, column=4, value=int(grp['Bestellmenge'].sum()))
            ws2.cell(row=row2, column=4).font      = summen_font2
            ws2.cell(row=row2, column=4).fill      = summen_fill2
            ws2.cell(row=row2, column=4).alignment = Alignment(horizontal='center')
            ws2.cell(row=row2, column=6, value=round(grp['Bestellwert'].sum(), 2))
            ws2.cell(row=row2, column=6).font          = summen_font2
            ws2.cell(row=row2, column=6).fill          = summen_fill2
            ws2.cell(row=row2, column=6).number_format = '#,##0.00'
            ws2.cell(row=row2, column=6).alignment     = Alignment(horizontal='right')
            ws2.cell(row=row2, column=3,
                     value=f'MBW: {mbw_val:.0f} € | fehlt: {fehlbetrag:.2f} €').font = Font(
                         italic=True, color=WEISS, name='Arial', size=9)
            ws2.cell(row=row2, column=3).fill = summen_fill2
            row2 += 2

        # Info-Banner Tab 2
        gesamt_unter = unter_mbw['Bestellwert'].sum()
        ws2.insert_rows(1)
        ws2.merge_cells('A1:K1')
        info2 = ws2['A1']
        info2.value = (
            f"Unter MBW – nicht bestellt  |  "
            f"Potentieller Bestellwert: {gesamt_unter:,.2f} EUR  |  "
            f"Hersteller: {unter_mbw['Hersteller'].nunique()}  |  "
            f"Positionen: {len(unter_mbw)}"
        )
        info2.font      = Font(bold=True, color=WEISS, name='Arial', size=11)
        info2.fill      = PatternFill('solid', fgColor='C0392B')
        info2.alignment = Alignment(horizontal='left', vertical='center')
        ws2.row_dimensions[1].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── Hauptworkflow ─────────────────────────────────────────────────────────────

def run_check(gmail, drive):
    """Verarbeitet wiederbestellung.xlsx falls sie im Projektordner liegt."""
    today = date.today()
    kw    = today.isocalendar()[1]
    year  = today.year

    if not os.path.exists(WIEDERBESTELLUNG_PATH):
        print("Keine wiederbestellung.xlsx gefunden — nichts zu tun.")
        return

    print(f"wiederbestellung.xlsx gefunden — starte Bestellprozedur KW{kw:02d}/{year}...")
    with open(WIEDERBESTELLUNG_PATH, 'rb') as f:
        excel_bytes = f.read()

    # Datei umbenennen damit sie nicht doppelt verarbeitet wird
    archiv_name = os.path.join(BASE_DIR, f'wiederbestellung_KW{kw:02d}_{year}_verarbeitet.xlsx')
    os.rename(WIEDERBESTELLUNG_PATH, archiv_name)
    print(f"Lokale Datei umbenannt: {os.path.basename(archiv_name)}")

    week_folder_id  = get_week_folder_id(drive, kw, year)
    stammdaten_id   = get_stammdaten_folder_id(drive)

    # Wiederbestelldatei in Drive archivieren
    upload_bytes_to_drive(
        drive, excel_bytes,
        f'wiederbestellung_KW{kw:02d}_{year}.xlsx',
        week_folder_id,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    # Letzte Bestellung aus lokalem purchase-orders Excel laden
    letzte_excel         = finde_letzte_bestellung_excel()
    letzte_bestellung_df = lade_letzte_bestellung(letzte_excel)

    # MBW-Ausnahmen aus Drive laden (optional)
    mbw_df        = download_csv_from_drive(drive, 'mbw_exceptions.csv', stammdaten_id)
    mbw_ausnahmen = dict(zip(mbw_df['Hersteller'], mbw_df['MBW'])) if mbw_df is not None else {}

    # Berechnung
    ergebnis = berechne_bestellvorschlag(excel_bytes, letzte_bestellung_df, mbw_ausnahmen)

    print("\nLOG:")
    for eintrag in ergebnis['log']:
        print(f"  {eintrag}")

    if ergebnis['bestellen'].empty:
        print("\nKeine Bestellungen über MBW.")
        return

    # Purchase-Order Excel lokal speichern
    excel_out  = erstelle_bestellsheet(ergebnis, kw, year)
    order_name = f'Purchase-Order-KW{kw:02d}-{year}.xlsx'
    order_path = os.path.join(BASE_DIR, order_name)
    with open(order_path, 'wb') as f:
        f.write(excel_out)
    print(f"Bestellvorschlag gespeichert: {order_name}")

    df_bestellen = ergebnis['bestellen']

    # Bestellhistorie lokal speichern (purchase-orders-DDMMYYYY.xlsx)
    historie_name, historie_pfad = speichere_bestellhistorie(ergebnis['df_input'], df_bestellen)

    # Beide Dateien in Drive archivieren
    upload_bytes_to_drive(
        drive, excel_out, order_name, week_folder_id,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    with open(historie_pfad, 'rb') as f:
        upload_bytes_to_drive(
            drive, f.read(), historie_name, week_folder_id,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    # E-Mail je Hersteller — aktivieren wenn Gmail autorisiert ist:
    # for hersteller, grp in df_bestellen.groupby('Hersteller'):
    #     send_order_email(gmail, hersteller, grp, kw, year,
    #                      grp['Bestellwert'].sum(), grp['MBW'].iloc[0])

    print(f"\nFertig. {df_bestellen['Hersteller'].nunique()} Hersteller, {len(df_bestellen)} Positionen.")

# ─── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Healthii Purchasing Agent – Leopold')
    parser.add_argument('--auth', action='store_true', help='OAuth einmalig autorisieren')
    args = parser.parse_args()

    gmail, drive = get_services()

    if args.auth:
        print("Authentifizierung erfolgreich.")
    else:
        run_check(gmail, drive)
